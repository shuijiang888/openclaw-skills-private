#!/usr/bin/env python3
"""Export Chenwei Excel workbooks to questionnaire.bundle.js and rubrics.bundle.js."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

COMPLIANCE_REPLACEMENTS = [
    (
        "提前准备竞争对手的负面信息，并在适当的时机传递给客户",
        "提前准备与竞争对手的客观对比要点与差异化证据，并在适当时机向客户清晰传递我司优势边界",
    ),
    (
        "在出现风险和问题时候，能够引导客户延迟开标或废标",
        "在出现风险和问题时，能够依据规则与客户沟通合理的程序调整或澄清机制",
    ),
]


def parse_score_cell(x) -> float | None:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    m = re.search(r"([\d.]+)", str(x))
    return float(m.group(1)) if m else None


def parse_questionnaire_sheet(ws) -> list[dict]:
    """Sales rep / manager / system ALL sheets: dim row has B=name and D–F scores; then numbered questions."""
    dims: list[dict] = []
    cur: dict | None = None
    opt_scores: dict | None = None
    for r in ws.iter_rows(values_only=True):
        if not r or len(r) < 4:
            continue
        b, c = r[1], r[2]
        d, e, f = r[3], r[4], r[5] if len(r) > 5 else (None, None, None)
        if b == "#" or (isinstance(b, str) and b.strip() == "#"):
            continue
        if isinstance(d, str) and str(d).strip() == "能够做到":
            continue
        if isinstance(b, str) and b.strip() and not re.match(r"^\d+$", b.strip()):
            if b.strip() in ("问题", "答案选项"):
                continue
            cur = {"name": b.strip(), "id": f"d_{len(dims)}", "questions": []}
            dims.append(cur)
            opt_scores = None
            # 管理者/销售：D–F 常为「4分」；营销体系：D–F 常为数值 2.5 / 2 / 1
            if d is not None and ("分" in str(d) or isinstance(d, (int, float))):
                opt_scores = {
                    "can": parse_score_cell(d),
                    "mostly": parse_score_cell(e),
                    "hard": parse_score_cell(f),
                }
            continue
        if cur is None:
            continue
        if isinstance(b, (int, float)) and float(b).is_integer() and isinstance(c, str) and c.strip():
            if not opt_scores or opt_scores.get("can") is None:
                continue
            cur["questions"].append(
                {
                    "id": f"q{len(cur['questions'])}",
                    "text": c.strip(),
                    "options": [
                        {"key": "can", "label": "能够做到", "score": opt_scores["can"]},
                        {"key": "mostly", "label": "基本做到", "score": opt_scores["mostly"]},
                        {"key": "hard", "label": "有难度", "score": opt_scores["hard"]},
                    ],
                }
            )
    return dims


def apply_compliance(obj) -> None:
    def walk(o):
        if isinstance(o, dict):
            for k, v in o.items():
                if k == "text" and isinstance(v, str):
                    for old, new in COMPLIANCE_REPLACEMENTS:
                        if old in v:
                            o[k] = v.replace(old, new)
                else:
                    walk(v)
        elif isinstance(o, list):
            for x in o:
                walk(x)

    walk(obj)


def attach_rep_max_points(dims: list[dict], mode_max: dict[str, float]) -> None:
    for d in dims:
        d["maxPoints"] = mode_max.get(d["name"])


def export_questionnaire(base: Path, out_dir: Path) -> dict:
    wb1 = openpyxl.load_workbook(base / "1.1能力测评问卷（销售人员） .xlsx", data_only=True)
    struct = wb1["问卷结构"]
    mode_cols = {3: "raw_material", 4: "semi_custom", 5: "standard_product", 6: "solution"}
    mode_max: dict[str, dict[str, float]] = {m: {} for m in mode_cols.values()}
    for row in struct.iter_rows(min_row=18, max_row=27, values_only=True):
        if not row[2] or not isinstance(row[2], str):
            continue
        dim_name = row[2].strip()
        if dim_name == "合计":
            continue
        for ci, mk in mode_cols.items():
            v = row[ci]
            if isinstance(v, (int, float)):
                mode_max[mk][dim_name] = float(v)
    sheet_map = {
        "raw_material": "原材料",
        "semi_custom": "半成品&定制件",
        "standard_product": "单品&标品",
        "solution": "解决方案",
    }
    bundle: dict = {"version": 1, "roles": {"sales_rep": {"modes": {}}}}
    for mk, sname in sheet_map.items():
        dims = parse_questionnaire_sheet(wb1[sname])
        attach_rep_max_points(dims, mode_max[mk])
        bundle["roles"]["sales_rep"]["modes"][mk] = {"dimensions": dims}
    wb1.close()

    wb2 = openpyxl.load_workbook(base / "2.1能力测评问卷（销售管理者） .xlsx", data_only=True)
    dims_m = parse_questionnaire_sheet(wb2["ALL"])
    for row in wb2["问卷结构"].iter_rows(min_row=14, max_row=20, values_only=True):
        if row[2] and isinstance(row[2], str) and row[3] is not None and isinstance(row[3], (int, float)):
            name = row[2].strip()
            for d in dims_m:
                if d["name"] == name:
                    d["maxPoints"] = float(row[3])
    bundle["roles"]["sales_manager"] = {"dimensions": dims_m}
    wb2.close()

    wb3 = openpyxl.load_workbook(base / "3.1能力测评问卷（营销体系）.xlsx", data_only=True)
    dims_sys = parse_questionnaire_sheet(wb3["ALL"])
    for row in wb3["问卷结构"].iter_rows(min_row=14, max_row=20, values_only=True):
        if row[2] and isinstance(row[2], str) and row[3] is not None and isinstance(row[3], (int, float)):
            name = row[2].strip()
            for d in dims_sys:
                if d["name"] == name:
                    d["maxPoints"] = float(row[3])
    bundle["roles"]["marketing_system"] = {"dimensions": dims_sys}
    wb3.close()

    apply_compliance(bundle)
    out = out_dir / "questionnaire.bundle.js"
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        f.write("window.UNIVERSAL_QUESTIONNAIRE = ")
        json.dump(bundle, f, ensure_ascii=False, indent=0)
        f.write(";\n")
    return bundle


def export_rubrics(base: Path, out_dir: Path) -> None:
    wb = openpyxl.load_workbook(base / "2.3测评结果数据库（营销管理者）.xlsx", data_only=True)
    ws = wb["销售管理者"]
    default_intro = ws.cell(5, 4).value
    if isinstance(default_intro, str):
        default_intro = default_intro.strip()
    else:
        default_intro = ""

    by_dimension: dict[str, dict] = {}
    for ri in range(6, ws.max_row + 1):
        name = ws.cell(ri, 3).value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        by_dimension[name] = {
            "C": {
                "evaluation": (ws.cell(ri, 4).value or "").strip() if ws.cell(ri, 4).value else "",
                "suggestion": (ws.cell(ri, 5).value or "").strip() if ws.cell(ri, 5).value else "",
                "training": (ws.cell(ri, 6).value or "").strip() if ws.cell(ri, 6).value else "",
            },
            "B": {
                "evaluation": (ws.cell(ri, 7).value or "").strip() if ws.cell(ri, 7).value else "",
                "suggestion": (ws.cell(ri, 8).value or "").strip() if ws.cell(ri, 8).value else "",
                "training": (ws.cell(ri, 9).value or "").strip() if ws.cell(ri, 9).value else "",
            },
            "A": {
                "evaluation": (ws.cell(ri, 10).value or "").strip() if ws.cell(ri, 10).value else "",
                "suggestion": "建议持续对标优秀实践，固化优势能力，并关注外部环境变化带来的新要求。",
                "training": "",
            },
        }

    courses_row = None
    t_ws = wb["培训课程汇总"]
    for r in t_ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if r[2] and str(r[2]).strip() == "培训等级":
            courses_row = r
            break
    training_catalog: dict[str, str] = {}
    if courses_row:
        # next row has 高级训战营
        idx = list(t_ws.iter_rows(min_row=1, max_row=5, values_only=True)).index(courses_row) + 2
        row = [t_ws.cell(idx, j).value for j in range(1, 6)]
        if row[2]:
            training_catalog[str(row[2]).strip()] = {
                "basic": (row[3] or "") if len(row) > 3 else "",
                "extended": (row[4] or "") if len(row) > 4 else "",
            }

    wb.close()

    generic = {
        "C": {
            "evaluation": "在「{dimension}」维度上，自评显示仍存在明显短板，关键动作落地不足，对业绩与协作可能形成制约。",
            "suggestion": "建议结合业务优先级制定改进计划：明确1–2个可度量目标，补齐流程与工具，并通过演练/复盘固化行为。",
            "training": "可参考《销售管理能力提升》《LTC 与销售过程管理》等通用课程（以贵司培训体系为准）。",
        },
        "B": {
            "evaluation": "在「{dimension}」维度上，自评显示已具备基础能力，但在一致性、深度与闭环上仍有提升空间。",
            "suggestion": "建议在既有框架上细化标准与检查点，强化过程数据与例会机制，推动从“会做”到“稳定产出”。",
            "training": "可结合岗位分层培训与训战营，针对薄弱子项做专项突破。",
        },
        "A": {
            "evaluation": "在「{dimension}」维度上，自评显示整体表现较为成熟，能够支撑当前业务要求。",
            "suggestion": "建议持续对标优秀实践，沉淀方法论，并在团队复制与赋能上投入精力。",
            "training": "",
        },
    }

    rubric_bundle = {
        "version": 1,
        "gradeBands": {"C": [30, 59], "B": [60, 79], "A": [80, 100]},
        "defaultIntroTemplate": default_intro,
        "sales_manager_by_dimension": by_dimension,
        "generic_by_band": generic,
        "training_catalog": training_catalog,
    }

    out = out_dir / "rubrics.bundle.js"
    with open(out, "w", encoding="utf-8") as f:
        f.write("window.UNIVERSAL_RUBRICS = ")
        json.dump(rubric_bundle, f, ensure_ascii=False, indent=0)
        f.write(";\n")


def main() -> int:
    base = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "陈玮营销诊断系统"
    out_dir = Path(__file__).resolve().parent.parent / "data"
    if not (base / "1.1能力测评问卷（销售人员） .xlsx").exists():
        print("Missing Excel folder:", base, file=sys.stderr)
        return 1
    export_questionnaire(base, out_dir)
    export_rubrics(base, out_dir)
    print("OK ->", out_dir)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
